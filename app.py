#!/usr/bin/env python3
"""
SWING TRADE DASHBOARD — Automated Live Data Engine
====================================================
Fetches live NASDAQ data, builds a formatted .xlsx, emails it on schedule,
and serves a web dashboard at http://localhost:5000

Schedule (CST): 7:00 AM | 9:30 AM | 12:00 PM | 2:45 PM

Usage:
  python app.py                  # Run full app (scheduler + web dashboard)
  python app.py --run-once       # Generate spreadsheet once and email it
  python app.py --web-only       # Web dashboard only (no scheduled emails)
  python app.py --no-email       # Scheduler runs but skips email (for testing)
"""

import os, sys, json, time, logging, argparse, threading, smtplib, io
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from pathlib import Path
from zoneinfo import ZoneInfo

# ---------------------------------------------------------------------------
# Third-party (installed by requirements.txt)
# ---------------------------------------------------------------------------
import yfinance as yf
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, render_template_string, send_file, jsonify
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import requests

# ---------------------------------------------------------------------------
# CONFIG  (override with environment variables or .env file)
# ---------------------------------------------------------------------------
CONFIG = {
    "EMAIL_TO":       os.getenv("EMAIL_TO", "el.capitan.44@gmail.com"),
    "SMTP_SERVER":    os.getenv("SMTP_SERVER", "smtp.gmail.com"),
    "SMTP_PORT":      int(os.getenv("SMTP_PORT", "587")),
    "SMTP_USER":      os.getenv("SMTP_USER", ""),        # your Gmail address
    "SMTP_PASSWORD":  os.getenv("SMTP_PASSWORD", ""),     # Gmail App Password
    "OUTPUT_DIR":     os.getenv("OUTPUT_DIR", "output"),
    "PORT":           int(os.getenv("PORT", "5000")),
    "TZ":             "America/Chicago",                   # CST
}

# The 10 NASDAQ tickers we track (market cap > $1B, high swing potential)
TICKERS = ["NVDA", "MSTR", "PLTR", "TSLA", "AMD", "AVGO", "PANW", "CRWD", "MU", "NFLX"]

SECTORS = {
    "NVDA": "Semiconductors", "MSTR": "Crypto / Software", "PLTR": "AI / Software",
    "TSLA": "EV / Automotive", "AMD": "Semiconductors", "AVGO": "Semiconductors",
    "PANW": "Cybersecurity", "CRWD": "Cybersecurity", "MU": "Semiconductors",
    "NFLX": "Streaming / Media",
}

CST = ZoneInfo(CONFIG["TZ"])
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("SwingBot")

# Global state: latest generated file path + summary dict
latest = {"file": None, "summary": {}, "generated_at": None, "error": None}

# ===========================================================================
#  DATA FETCHING
# ===========================================================================

def fetch_stock_data() -> pd.DataFrame:
    """Pull live/recent price data for all tickers via yfinance."""
    log.info("Fetching stock data for %s", TICKERS)
    rows = []
    three_months_ago = datetime.now() - timedelta(days=90)

    for ticker in TICKERS:
        try:
            tk = yf.Ticker(ticker)
            info = tk.info or {}

            current = info.get("currentPrice") or info.get("regularMarketPrice") or info.get("previousClose", 0)
            prev_close = info.get("previousClose") or info.get("regularMarketPreviousClose", current)

            # 3-month-ago price from historical data
            hist = tk.history(start=three_months_ago.strftime("%Y-%m-%d"),
                             end=(three_months_ago + timedelta(days=5)).strftime("%Y-%m-%d"))
            price_3mo = float(hist["Close"].iloc[0]) if not hist.empty else current * 0.85

            mkt_cap = info.get("marketCap", 0) / 1e9  # billions
            avg_vol = info.get("averageVolume", 0) / 1e6  # millions
            beta = info.get("beta", 1.0) or 1.0
            fifty_two_high = info.get("fiftyTwoWeekHigh", current)
            fifty_two_low = info.get("fiftyTwoWeekLow", current)
            day_high = info.get("dayHigh", current)
            day_low = info.get("dayLow", current)

            daily_pct = ((current - prev_close) / prev_close) if prev_close else 0
            three_mo_pct = ((current - price_3mo) / price_3mo) if price_3mo else 0

            # --- Swing Score Calculation ---
            # Volatility (0-35): based on beta, 52-wk range, intraday range
            range_pct = (fifty_two_high - fifty_two_low) / fifty_two_low if fifty_two_low else 0
            intraday_pct = (day_high - day_low) / day_low if day_low else 0
            vol_score = min(35, int(beta * 8 + range_pct * 10 + intraday_pct * 200))

            # Momentum (0-35): based on 3-month return magnitude + daily move
            mom_score = min(35, int(abs(three_mo_pct) * 40 + abs(daily_pct) * 150))

            # Volume (0-30): based on avg daily volume (higher = better liquidity)
            liq_score = min(30, int(avg_vol * 0.3 + 5)) if avg_vol > 0 else 5

            swing_score = vol_score + mom_score + liq_score

            rows.append({
                "Ticker": ticker, "Company": info.get("shortName", ticker),
                "Sector": SECTORS.get(ticker, "N/A"),
                "Current Price": round(current, 2),
                "Prev Close": round(prev_close, 2),
                "3-Mo Ago Price": round(price_3mo, 2),
                "Daily % Chg": round(daily_pct, 4),
                "3-Month % Chg": round(three_mo_pct, 4),
                "Market Cap ($B)": round(mkt_cap, 1),
                "Avg Vol (M)": round(avg_vol, 1),
                "Beta": round(beta, 2),
                "Swing Score": min(swing_score, 100),
                "Vol Score": vol_score,
                "Mom Score": mom_score,
                "Liq Score": liq_score,
                "52wk High": round(fifty_two_high, 2),
                "52wk Low": round(fifty_two_low, 2),
            })
        except Exception as e:
            log.warning("Error fetching %s: %s", ticker, e)
            rows.append({"Ticker": ticker, "Company": ticker, "Sector": SECTORS.get(ticker, ""),
                         "Current Price": 0, "Prev Close": 0, "3-Mo Ago Price": 0,
                         "Daily % Chg": 0, "3-Month % Chg": 0,
                         "Market Cap ($B)": 0, "Avg Vol (M)": 0, "Beta": 1,
                         "Swing Score": 0, "Vol Score": 0, "Mom Score": 0, "Liq Score": 0,
                         "52wk High": 0, "52wk Low": 0})

    df = pd.DataFrame(rows).sort_values("Swing Score", ascending=False).reset_index(drop=True)
    df.insert(0, "Rank", range(1, len(df) + 1))
    log.info("Data fetched for %d tickers", len(df))
    return df


def fetch_futures_data() -> list[dict]:
    """Fetch index futures / macro data via yfinance."""
    log.info("Fetching futures & macro data")
    symbols = {
        "ES=F": "S&P 500 Futures",
        "NQ=F": "Nasdaq 100 Futures",
        "YM=F": "Dow Futures",
        "RTY=F": "Russell 2000 Futures",
        "^VIX": "VIX (Fear Index)",
        "^TNX": "10-Year Treasury Yield",
        "DX-Y.NYB": "US Dollar Index (DXY)",
    }
    results = []
    for sym, name in symbols.items():
        try:
            tk = yf.Ticker(sym)
            info = tk.info or {}
            current = info.get("regularMarketPrice") or info.get("previousClose", 0)
            prev = info.get("regularMarketPreviousClose") or info.get("previousClose", current)
            change = current - prev
            pct = (change / prev * 100) if prev else 0

            if "VIX" in name:
                signal = "DECREASING FEAR" if change < 0 else "INCREASING FEAR"
            elif pct > 0.3:
                signal = "BULLISH"
            elif pct > 0:
                signal = "SLIGHTLY BULLISH"
            elif pct > -0.3:
                signal = "NEUTRAL"
            else:
                signal = "BEARISH"

            results.append({
                "name": name, "level": round(current, 2),
                "change": round(change, 2), "pct": round(pct, 2), "signal": signal,
            })
        except Exception as e:
            log.warning("Futures fetch error %s: %s", sym, e)
            results.append({"name": name, "level": 0, "change": 0, "pct": 0, "signal": "N/A"})
    return results


def fetch_market_news_context() -> dict:
    """Return static/semi-static macro context that gets embedded in the sheet."""
    # In a production version, you'd scrape RSS feeds or use a news API.
    # Here we provide a structure that you can enhance later.
    return {
        "fed_rate": "3.50% - 3.75%",
        "fed_status": "Rates held steady at Jan 27-28 meeting. 2 dissents favored a cut.",
        "next_fomc": "March 17-18, 2026",
        "market_expects": "93% probability rates held in March; ~65bps of cuts priced for 2026",
        "overall_market": "BULL MARKET — S&P 500 near all-time highs",
    }


# ===========================================================================
#  SPREADSHEET GENERATION
# ===========================================================================

def build_spreadsheet(df: pd.DataFrame, futures: list, macro: dict) -> str:
    """Generate the formatted .xlsx and return its file path."""
    wb = Workbook()
    now = datetime.now(CST)
    timestamp = now.strftime("%Y%m%d_%H%M")

    # --- Styles ---
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    sub_fill = PatternFill("solid", fgColor="D6E4F0")
    dfont = Font(name="Arial", size=10)
    bfont = Font(name="Arial", bold=True, size=10)
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    green_font = Font(name="Arial", size=10, color="006100")
    red_font = Font(name="Arial", size=10, color="C00000")
    bdr = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    def style_cell(ws, r, c, val, font=dfont, fill=None, align="center", fmt=None):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = font
        cell.border = bdr
        cell.alignment = Alignment(horizontal=align, wrap_text=True)
        if fill: cell.fill = fill
        if fmt: cell.number_format = fmt
        return cell

    # ============================
    # SHEET 1: MARKET OVERVIEW
    # ============================
    ws1 = wb.active
    ws1.title = "Market Overview"
    ws1.merge_cells("A1:G1")
    ws1["A1"] = "SWING TRADE MARKET INTELLIGENCE DASHBOARD"
    ws1["A1"].font = Font(name="Arial", bold=True, size=16, color="1F4E79")
    ws1["A1"].alignment = Alignment(horizontal="center")
    ws1.merge_cells("A2:G2")
    ws1["A2"] = f"Generated: {now.strftime('%B %d, %Y at %I:%M %p %Z')}"
    ws1["A2"].font = Font(name="Arial", italic=True, size=10, color="666666")
    ws1["A2"].alignment = Alignment(horizontal="center")

    # -- Futures table --
    r = 4
    ws1.merge_cells(f"A{r}:G{r}")
    c = ws1[f"A{r}"]
    c.value = "PRE-MARKET / CURRENT FUTURES & MACRO INDICATORS"
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2E75B6")
    c.alignment = Alignment(horizontal="center")

    r = 5
    for i, h in enumerate(["Index / Future", "Level", "Change", "% Change", "Signal", "", ""], 1):
        style_cell(ws1, r, i, h, font=hdr_font, fill=hdr_fill)

    for idx, fut in enumerate(futures):
        row = 6 + idx
        style_cell(ws1, row, 1, fut["name"], align="left")
        style_cell(ws1, row, 2, fut["level"], fmt="#,##0.00")
        ch = fut["change"]
        style_cell(ws1, row, 3, f"{'+' if ch>=0 else ''}{ch}", font=green_font if ch>=0 else red_font)
        style_cell(ws1, row, 4, f"{'+' if fut['pct']>=0 else ''}{fut['pct']}%",
                   font=green_font if fut["pct"]>=0 else red_font)
        sig = fut["signal"]
        sig_fill = green_fill if "BULL" in sig or "DECREASING" in sig else (red_fill if "BEAR" in sig or "INCREASING" in sig else sub_fill)
        style_cell(ws1, row, 5, sig, font=bfont, fill=sig_fill)

    # Futures verdict
    bull_count = sum(1 for f in futures if "BULL" in f["signal"] or "DECREASING" in f["signal"])
    verdict = "BULLISH" if bull_count >= 4 else ("MIXED" if bull_count >= 2 else "BEARISH")
    vr = 6 + len(futures)
    ws1.merge_cells(f"A{vr}:G{vr}")
    ws1[f"A{vr}"] = f"⚡ FUTURES VERDICT: {verdict} ({bull_count}/{len(futures)} indicators positive). Plan swing entries accordingly."
    ws1[f"A{vr}"].font = Font(name="Arial", bold=True, size=10,
                                color="006100" if verdict == "BULLISH" else ("C00000" if verdict == "BEARISH" else "BF8F00"))
    ws1[f"A{vr}"].fill = green_fill if verdict == "BULLISH" else (red_fill if verdict == "BEARISH" else sub_fill)

    # -- Overall Market + Fed Section --
    r = vr + 2
    ws1.merge_cells(f"A{r}:G{r}")
    c = ws1[f"A{r}"]
    c.value = "OVERALL MARKET & FEDERAL RESERVE STATUS"
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="C00000")
    c.alignment = Alignment(horizontal="center")

    r += 1
    for i, h in enumerate(["Item", "Value", "Status", "", "", "", ""], 1):
        style_cell(ws1, r, i, h, font=hdr_font, fill=PatternFill("solid", fgColor="8B0000"))

    macro_rows = [
        ["Overall Market", macro["overall_market"], "BULL"],
        ["Fed Funds Rate", macro["fed_rate"], "HELD"],
        ["Last FOMC Decision", macro["fed_status"], "DIVIDED"],
        ["Next FOMC Meeting", macro["next_fomc"], "UPCOMING"],
        ["Rate Cut Expectations", macro["market_expects"], "DOVISH LEAN"],
    ]
    for idx, mr in enumerate(macro_rows):
        row = r + 1 + idx
        style_cell(ws1, row, 1, mr[0], align="left")
        style_cell(ws1, row, 2, mr[1], align="left")
        st = mr[2]
        fill = green_fill if st in ("BULL","DOVISH LEAN") else (red_fill if st == "BEAR" else sub_fill)
        style_cell(ws1, row, 3, st, font=bfont, fill=fill)

    for col, w in enumerate([35, 28, 25, 20, 35, 45, 5], 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ============================
    # SHEET 2: TOP 10 SWING TRADES
    # ============================
    ws2 = wb.create_sheet("Top 10 Swing Trades")
    ws2.merge_cells("A1:P1")
    ws2["A1"] = "TOP 10 NASDAQ SWING TRADE OPPORTUNITIES (Market Cap > $1B)"
    ws2["A1"].font = Font(name="Arial", bold=True, size=16, color="1F4E79")
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.merge_cells("A2:P2")
    ws2["A2"] = f"Updated: {now.strftime('%B %d, %Y at %I:%M %p %Z')} | Top 3 highlighted in yellow"
    ws2["A2"].font = Font(name="Arial", italic=True, size=10, color="666666")
    ws2["A2"].alignment = Alignment(horizontal="center")

    cols = ["Rank","Ticker","Company","Sector","Current Price","Prev Close",
            "3-Mo Ago Price","Daily % Chg","3-Month % Chg","Market Cap ($B)",
            "Avg Vol (M)","Swing Score","Vol Score","Mom Score","Liq Score","52wk Range"]
    for i, h in enumerate(cols, 1):
        style_cell(ws2, 3, i, h, font=hdr_font, fill=hdr_fill)

    for idx, (_, row_data) in enumerate(df.iterrows()):
        r = 4 + idx
        is_top3 = row_data["Rank"] <= 3
        row_fill = yellow_fill if is_top3 else None

        style_cell(ws2, r, 1, int(row_data["Rank"]), fill=row_fill)
        style_cell(ws2, r, 2, row_data["Ticker"], font=bfont, fill=row_fill)
        style_cell(ws2, r, 3, row_data["Company"], align="left", fill=row_fill)
        style_cell(ws2, r, 4, row_data["Sector"], fill=row_fill)

        style_cell(ws2, r, 5, row_data["Current Price"], fmt="$#,##0.00", fill=row_fill)
        style_cell(ws2, r, 6, row_data["Prev Close"], fmt="$#,##0.00", fill=row_fill)
        style_cell(ws2, r, 7, row_data["3-Mo Ago Price"], fmt="$#,##0.00", fill=row_fill)

        # Daily % change - color code
        dpct = row_data["Daily % Chg"]
        d_font = Font(name="Arial", bold=True, size=10, color="006100" if dpct >= 0 else "C00000")
        d_fill = green_fill if dpct >= 0 else red_fill
        if is_top3:
            # keep yellow but tint font
            d_fill = yellow_fill
        style_cell(ws2, r, 8, dpct, font=d_font, fill=d_fill, fmt="0.00%")

        mpct = row_data["3-Month % Chg"]
        m_font = Font(name="Arial", bold=True, size=10, color="006100" if mpct >= 0 else "C00000")
        m_fill = green_fill if mpct >= 0 else red_fill
        if is_top3:
            m_fill = yellow_fill
        style_cell(ws2, r, 9, mpct, font=m_font, fill=m_fill, fmt="0.00%")

        style_cell(ws2, r, 10, row_data["Market Cap ($B)"], fmt="#,##0.0", fill=row_fill)
        style_cell(ws2, r, 11, row_data["Avg Vol (M)"], fmt="#,##0.0", fill=row_fill)

        # Swing Score - bold, colored
        ss = row_data["Swing Score"]
        ss_fill = green_fill if ss >= 80 else (sub_fill if ss >= 60 else red_fill)
        if is_top3: ss_fill = yellow_fill
        style_cell(ws2, r, 12, ss, font=bfont, fill=ss_fill)
        style_cell(ws2, r, 13, row_data["Vol Score"], fill=row_fill)
        style_cell(ws2, r, 14, row_data["Mom Score"], fill=row_fill)
        style_cell(ws2, r, 15, row_data["Liq Score"], fill=row_fill)

        rng = f"${row_data['52wk Low']:.2f} – ${row_data['52wk High']:.2f}"
        style_cell(ws2, r, 16, rng, fill=row_fill)

    for col, w in enumerate([6,8,28,18,13,13,13,12,13,14,12,12,10,10,10,22], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # Save
    os.makedirs(CONFIG["OUTPUT_DIR"], exist_ok=True)
    filename = f"swing_dashboard_{timestamp}.xlsx"
    filepath = os.path.join(CONFIG["OUTPUT_DIR"], filename)
    wb.save(filepath)
    log.info("Spreadsheet saved: %s", filepath)

    # Also save as 'latest.xlsx' for the web route
    latest_path = os.path.join(CONFIG["OUTPUT_DIR"], "latest.xlsx")
    wb.save(latest_path)

    return filepath


# ===========================================================================
#  EMAIL
# ===========================================================================

def send_email(filepath: str, skip=False):
    """Email the spreadsheet as an attachment via SMTP / Gmail."""
    if skip:
        log.info("Email skipped (--no-email flag)")
        return
    user = CONFIG["SMTP_USER"]
    pwd  = CONFIG["SMTP_PASSWORD"]
    if not user or not pwd:
        log.warning("SMTP credentials not set — skipping email. Set SMTP_USER and SMTP_PASSWORD.")
        return

    msg = MIMEMultipart()
    msg["From"] = user
    msg["To"] = CONFIG["EMAIL_TO"]
    now_str = datetime.now(CST).strftime("%I:%M %p CST on %b %d, %Y")
    msg["Subject"] = f"🔔 Swing Trade Dashboard — {now_str}"

    body = (
        f"Your Swing Trade Dashboard has been refreshed as of {now_str}.\n\n"
        "The attached spreadsheet contains:\n"
        "• Pre-market futures & macro indicators\n"
        "• Overall market bull/bear status\n"
        "• Sector breakdown\n"
        "• Top 10 NASDAQ swing trade opportunities with Swing Scores\n"
        "• Fed/FOMC status & upcoming events\n\n"
        "— Swing Trade Bot 🤖"
    )
    msg.attach(MIMEText(body, "plain"))

    with open(filepath, "rb") as f:
        part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(filepath)}")
        msg.attach(part)

    try:
        with smtplib.SMTP(CONFIG["SMTP_SERVER"], CONFIG["SMTP_PORT"]) as server:
            server.ehlo()
            server.starttls()
            server.login(user, pwd)
            server.sendmail(user, CONFIG["EMAIL_TO"], msg.as_string())
        log.info("Email sent to %s", CONFIG["EMAIL_TO"])
    except Exception as e:
        log.error("Email failed: %s", e)


# ===========================================================================
#  MAIN JOB (called by scheduler and on-demand)
# ===========================================================================

def run_job(skip_email=False):
    """Full pipeline: fetch → build → email."""
    log.info("=" * 60)
    log.info("RUNNING SWING TRADE DASHBOARD JOB")
    log.info("=" * 60)
    try:
        df = fetch_stock_data()
        futures = fetch_futures_data()
        macro = fetch_market_news_context()
        filepath = build_spreadsheet(df, futures, macro)
        send_email(filepath, skip=skip_email)

        latest["file"] = filepath
        latest["generated_at"] = datetime.now(CST).isoformat()
        latest["summary"] = {
            "top3": df.head(3)[["Ticker", "Swing Score", "Daily % Chg", "Current Price"]].to_dict("records"),
            "futures_verdict": "BULLISH" if sum(1 for f in futures if "BULL" in f["signal"] or "DECREASING" in f["signal"]) >= 4 else "MIXED",
            "futures": futures,
        }
        latest["error"] = None
        log.info("Job complete. File: %s", filepath)
    except Exception as e:
        log.error("Job failed: %s", e, exc_info=True)
        latest["error"] = str(e)


# ===========================================================================
#  WEB DASHBOARD  (Flask)
# ===========================================================================

app = Flask(__name__)

DASHBOARD_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Swing Trade Dashboard</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'Segoe UI', Arial, sans-serif; background: #0d1117; color: #e6edf3; }
  .header { background: linear-gradient(135deg, #1f4e79 0%, #2e75b6 100%);
             padding: 24px; text-align: center; }
  .header h1 { font-size: 28px; margin-bottom: 4px; }
  .header p { opacity: 0.8; font-size: 14px; }
  .container { max-width: 1400px; margin: 20px auto; padding: 0 16px; }
  .card { background: #161b22; border: 1px solid #30363d; border-radius: 8px;
           padding: 20px; margin-bottom: 16px; }
  .card h2 { color: #58a6ff; margin-bottom: 12px; font-size: 18px; }
  .grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); gap: 16px; }
  table { width: 100%; border-collapse: collapse; font-size: 14px; }
  th { background: #1f4e79; color: white; padding: 10px 8px; text-align: center; }
  td { padding: 8px; text-align: center; border-bottom: 1px solid #30363d; }
  .pos { color: #3fb950; font-weight: bold; }
  .neg { color: #f85149; font-weight: bold; }
  .top3 { background: rgba(255,255,0,0.12); }
  .badge { display: inline-block; padding: 3px 10px; border-radius: 12px; font-size: 12px; font-weight: bold; }
  .bull { background: #0d4429; color: #3fb950; }
  .bear { background: #4a1520; color: #f85149; }
  .neutral { background: #3d2e00; color: #d29922; }
  .btn { display: inline-block; background: #2e75b6; color: white; padding: 10px 24px;
         border-radius: 6px; text-decoration: none; font-weight: bold; margin: 4px; }
  .btn:hover { background: #1f4e79; }
  .refresh-info { text-align: center; padding: 10px; color: #8b949e; font-size: 13px; }
  .actions { text-align: center; margin: 16px 0; }
  #status { text-align: center; padding: 8px; color: #d29922; }
</style>
</head>
<body>
<div class="header">
  <h1>⚡ Swing Trade Dashboard</h1>
  <p>Live NASDAQ Swing Trade Intelligence — Auto-refreshes 4x daily (7am, 9:30am, 12pm, 2:45pm CST)</p>
</div>
<div class="container">
  <div class="actions">
    <a class="btn" href="/download">📥 Download Latest .xlsx</a>
    <a class="btn" href="/refresh" id="refreshBtn">🔄 Refresh Now</a>
  </div>
  <div id="status"></div>
  <div id="content">Loading dashboard data...</div>
</div>
<script>
async function loadDashboard() {
  try {
    const res = await fetch('/api/data');
    const data = await res.json();
    if (data.error) {
      document.getElementById('content').innerHTML = `<div class="card"><p>Error: ${data.error}. Click Refresh.</p></div>`;
      return;
    }
    let html = '';

    // Futures
    html += '<div class="card"><h2>📊 Futures & Macro Indicators</h2>';
    html += `<p style="margin-bottom:12px">Verdict: <span class="badge ${data.futures_verdict==='BULLISH'?'bull':'neutral'}">${data.futures_verdict}</span></p>`;
    html += '<table><tr><th>Index</th><th>Level</th><th>Change</th><th>%</th><th>Signal</th></tr>';
    (data.futures||[]).forEach(f => {
      const cls = f.pct >= 0 ? 'pos' : 'neg';
      const sig_cls = f.signal.includes('BULL') || f.signal.includes('DECREASING') ? 'bull' : (f.signal.includes('BEAR') || f.signal.includes('INCREASING') ? 'bear' : 'neutral');
      html += `<tr><td style="text-align:left">${f.name}</td><td>${f.level.toLocaleString()}</td>
               <td class="${cls}">${f.change>=0?'+':''}${f.change}</td>
               <td class="${cls}">${f.pct>=0?'+':''}${f.pct}%</td>
               <td><span class="badge ${sig_cls}">${f.signal}</span></td></tr>`;
    });
    html += '</table></div>';

    // Top 10
    html += '<div class="card"><h2>🏆 Top 10 Swing Trade Opportunities</h2>';
    html += '<table><tr><th>#</th><th>Ticker</th><th>Price</th><th>Daily %</th><th>3-Mo %</th><th>Mkt Cap</th><th>Swing Score</th></tr>';
    (data.stocks||[]).forEach((s, i) => {
      const tr_cls = i < 3 ? ' class="top3"' : '';
      const d_cls = s.daily_pct >= 0 ? 'pos' : 'neg';
      const m_cls = s.three_mo_pct >= 0 ? 'pos' : 'neg';
      html += `<tr${tr_cls}><td>${i+1}</td><td><strong>${s.ticker}</strong></td>
               <td>$${s.price.toFixed(2)}</td>
               <td class="${d_cls}">${(s.daily_pct*100).toFixed(2)}%</td>
               <td class="${m_cls}">${(s.three_mo_pct*100).toFixed(2)}%</td>
               <td>$${s.mkt_cap.toFixed(1)}B</td>
               <td><strong>${s.swing_score}</strong></td></tr>`;
    });
    html += '</table></div>';

    html += `<div class="refresh-info">Last updated: ${data.generated_at || 'N/A'}</div>`;
    document.getElementById('content').innerHTML = html;
  } catch(e) {
    document.getElementById('content').innerHTML = '<div class="card"><p>Could not load data. Click Refresh to generate.</p></div>';
  }
}
document.getElementById('refreshBtn').addEventListener('click', async (e) => {
  e.preventDefault();
  document.getElementById('status').innerText = '⏳ Refreshing data... this takes 30-60 seconds...';
  const res = await fetch('/refresh');
  const data = await res.json();
  document.getElementById('status').innerText = data.status === 'ok' ? '✅ Refresh complete!' : '❌ ' + data.error;
  setTimeout(loadDashboard, 1000);
});
loadDashboard();
setInterval(loadDashboard, 300000); // auto-refresh page every 5 min
</script>
</body>
</html>
"""

@app.route("/")
def index():
    return render_template_string(DASHBOARD_HTML)

@app.route("/api/data")
def api_data():
    if latest["error"] and not latest["file"]:
        return jsonify({"error": latest["error"]})
    if not latest["file"]:
        return jsonify({"error": "No data yet. Click Refresh to generate."})

    summary = latest["summary"]
    # Build stocks list from latest file
    stocks = []
    try:
        df = pd.read_excel(os.path.join(CONFIG["OUTPUT_DIR"], "latest.xlsx"),
                           sheet_name="Top 10 Swing Trades", skiprows=2)
        for _, r in df.iterrows():
            stocks.append({
                "ticker": r.get("Ticker", ""),
                "price": float(r.get("Current Price", 0)),
                "daily_pct": float(r.get("Daily % Chg", 0)),
                "three_mo_pct": float(r.get("3-Month % Chg", 0)),
                "mkt_cap": float(r.get("Market Cap ($B)", 0)),
                "swing_score": int(r.get("Swing Score", 0)),
            })
    except Exception:
        stocks = []

    return jsonify({
        "generated_at": latest["generated_at"],
        "futures_verdict": summary.get("futures_verdict", "N/A"),
        "futures": summary.get("futures", []),
        "stocks": stocks,
    })

@app.route("/refresh")
def refresh():
    try:
        skip = "--no-email" in sys.argv
        run_job(skip_email=skip)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)})

@app.route("/download")
def download():
    path = os.path.join(CONFIG["OUTPUT_DIR"], "latest.xlsx")
    if os.path.exists(path):
        return send_file(path, as_attachment=True,
                         download_name=f"swing_dashboard_{datetime.now(CST).strftime('%Y%m%d')}.xlsx")
    return "No file generated yet. Visit /refresh first.", 404


# ===========================================================================
#  SCHEDULER
# ===========================================================================

def start_scheduler(skip_email=False):
    """Schedule jobs at 7:00, 9:30, 12:00, 2:45 CST."""
    sched = BackgroundScheduler(timezone=CONFIG["TZ"])
    job_fn = lambda: run_job(skip_email=skip_email)

    sched.add_job(job_fn, CronTrigger(hour=7,  minute=0),  id="morning",   name="7:00 AM CST")
    sched.add_job(job_fn, CronTrigger(hour=9,  minute=30), id="premarket",  name="9:30 AM CST")
    sched.add_job(job_fn, CronTrigger(hour=12, minute=0),  id="midday",    name="12:00 PM CST")
    sched.add_job(job_fn, CronTrigger(hour=14, minute=45), id="afternoon", name="2:45 PM CST")

    sched.start()
    log.info("Scheduler started. Jobs: 7:00, 9:30, 12:00, 14:45 CST")
    return sched


# ===========================================================================
#  ENTRY POINT
# ===========================================================================

def main():
    parser = argparse.ArgumentParser(description="Swing Trade Dashboard Bot")
    parser.add_argument("--run-once", action="store_true", help="Generate spreadsheet once and email, then exit")
    parser.add_argument("--web-only", action="store_true", help="Run web dashboard only, no scheduler")
    parser.add_argument("--no-email", action="store_true", help="Skip email sending (for testing)")
    args = parser.parse_args()

    if args.run_once:
        run_job(skip_email=args.no_email)
        return

    # Run an initial job immediately so the dashboard has data
    log.info("Running initial data fetch...")
    threading.Thread(target=run_job, kwargs={"skip_email": args.no_email}, daemon=True).start()

    if not args.web_only:
        start_scheduler(skip_email=args.no_email)

    log.info("Starting web dashboard on port %d", CONFIG["PORT"])
    app.run(host="0.0.0.0", port=CONFIG["PORT"], debug=False)


if __name__ == "__main__":
    main()
