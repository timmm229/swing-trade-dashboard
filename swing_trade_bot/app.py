#!/usr/bin/env python3
"""
SWING TRADE DASHBOARD v2.0 — Automated Live Data Engine + Buy Signals
======================================================================
Fetches live NASDAQ data + technical indicators, builds a formatted .xlsx,
emails it on schedule, and serves a web dashboard.

Sheets:
  1. Market Overview — Futures, macro, sector status, Fed events
  2. Top 10 Swing Trades — Ranked by Swing Score
  3. Buy Signals — Technical analysis with BUY/WATCH/AVOID ratings
  4. Trading Notes — Catalysts, risk factors, disclaimer

Schedule (CST): 7:00 AM | 9:30 AM | 12:00 PM | 2:45 PM

Usage:
  python app.py                  # Full app (scheduler + web + email)
  python app.py --run-once       # Generate once, email, exit
  python app.py --no-email       # Skip email (testing)
  python app.py --web-only       # Web dashboard only
"""

import os, sys, json, time, logging, argparse, threading, smtplib, io, math
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from pathlib import Path
from zoneinfo import ZoneInfo

import yfinance as yf
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.gridspec import GridSpec
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
from flask import Flask, render_template_string, send_file, jsonify
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------
CONFIG = {
    "EMAIL_TO":       os.getenv("EMAIL_TO", "el.capitan.44@gmail.com"),
    "SMTP_SERVER":    os.getenv("SMTP_SERVER", "smtp.gmail.com"),
    "SMTP_PORT":      int(os.getenv("SMTP_PORT", "587")),
    "SMTP_USER":      os.getenv("SMTP_USER", ""),
    "SMTP_PASSWORD":  os.getenv("SMTP_PASSWORD", ""),
    "OUTPUT_DIR":     os.getenv("OUTPUT_DIR", "output"),
    "PORT":           int(os.getenv("PORT", "5000")),
    "TZ":             "America/Chicago",
}

TICKERS = ["NVDA", "MSTR", "PLTR", "TSLA", "AMD", "AVGO", "PANW", "CRWD", "MU", "NFLX"]

SECTORS = {
    "NVDA": "Semiconductors", "MSTR": "Crypto / Software", "PLTR": "AI / Software",
    "TSLA": "EV / Automotive", "AMD": "Semiconductors", "AVGO": "Semiconductors",
    "PANW": "Cybersecurity", "CRWD": "Cybersecurity", "MU": "Semiconductors",
    "NFLX": "Streaming / Media",
}

CST = ZoneInfo(CONFIG["TZ"])
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("SwingBot")

latest = {"file": None, "summary": {}, "generated_at": None, "error": None, "buy_signals": [], "chart_images": {}}


# ===========================================================================
#  TECHNICAL INDICATOR CALCULATIONS
# ===========================================================================

def calc_rsi(series, period=14):
    delta = series.diff()
    gain = delta.where(delta > 0, 0.0)
    loss = (-delta).where(delta < 0, 0.0)
    avg_gain = gain.rolling(window=period, min_periods=period).mean()
    avg_loss = loss.rolling(window=period, min_periods=period).mean()
    rs = avg_gain / avg_loss.replace(0, np.nan)
    rsi = 100 - (100 / (1 + rs))
    return rsi


def calc_macd(series, fast=12, slow=26, signal=9):
    ema_fast = series.ewm(span=fast, adjust=False).mean()
    ema_slow = series.ewm(span=slow, adjust=False).mean()
    macd_line = ema_fast - ema_slow
    signal_line = macd_line.ewm(span=signal, adjust=False).mean()
    histogram = macd_line - signal_line
    return macd_line, signal_line, histogram


def calc_bollinger(series, period=20, std_dev=2):
    sma = series.rolling(window=period).mean()
    std = series.rolling(window=period).std()
    upper = sma + (std * std_dev)
    lower = sma - (std * std_dev)
    return upper, sma, lower


def calc_sma(series, period):
    return series.rolling(window=period).mean()


# ===========================================================================
#  DATA FETCHING
# ===========================================================================

def fetch_stock_data():
    log.info("Fetching stock data for %s", TICKERS)
    rows = []
    three_months_ago = datetime.now() - timedelta(days=90)

    for ticker in TICKERS:
        try:
            tk = yf.Ticker(ticker)
            info = tk.info or {}
            current = info.get("currentPrice") or info.get("regularMarketPrice") or info.get("previousClose", 0)
            prev_close = info.get("previousClose") or info.get("regularMarketPreviousClose", current)
            hist = tk.history(start=three_months_ago.strftime("%Y-%m-%d"),
                             end=(three_months_ago + timedelta(days=5)).strftime("%Y-%m-%d"))
            price_3mo = float(hist["Close"].iloc[0]) if not hist.empty else current * 0.85
            mkt_cap = info.get("marketCap", 0) / 1e9
            avg_vol = info.get("averageVolume", 0) / 1e6
            beta = info.get("beta", 1.0) or 1.0
            fifty_two_high = info.get("fiftyTwoWeekHigh", current)
            fifty_two_low = info.get("fiftyTwoWeekLow", current)
            day_high = info.get("dayHigh", current)
            day_low = info.get("dayLow", current)
            daily_pct = ((current - prev_close) / prev_close) if prev_close else 0
            three_mo_pct = ((current - price_3mo) / price_3mo) if price_3mo else 0

            range_pct = (fifty_two_high - fifty_two_low) / fifty_two_low if fifty_two_low else 0
            intraday_pct = (day_high - day_low) / day_low if day_low else 0
            vol_score = min(35, int(beta * 8 + range_pct * 10 + intraday_pct * 200))
            mom_score = min(35, int(abs(three_mo_pct) * 40 + abs(daily_pct) * 150))
            liq_score = min(30, int(avg_vol * 0.3 + 5)) if avg_vol > 0 else 5
            swing_score = vol_score + mom_score + liq_score

            rows.append({
                "Ticker": ticker, "Company": info.get("shortName", ticker),
                "Sector": SECTORS.get(ticker, "N/A"),
                "Current Price": round(current, 2), "Prev Close": round(prev_close, 2),
                "3-Mo Ago Price": round(price_3mo, 2),
                "Daily % Chg": round(daily_pct, 4), "3-Month % Chg": round(three_mo_pct, 4),
                "Market Cap ($B)": round(mkt_cap, 1), "Avg Vol (M)": round(avg_vol, 1),
                "Beta": round(beta, 2),
                "Swing Score": min(swing_score, 100),
                "Vol Score": vol_score, "Mom Score": mom_score, "Liq Score": liq_score,
                "52wk High": round(fifty_two_high, 2), "52wk Low": round(fifty_two_low, 2),
            })
        except Exception as e:
            log.warning("Error fetching %s: %s", ticker, e)
            rows.append({"Ticker": ticker, "Company": ticker, "Sector": SECTORS.get(ticker, ""),
                         "Current Price": 0, "Prev Close": 0, "3-Mo Ago Price": 0,
                         "Daily % Chg": 0, "3-Month % Chg": 0,
                         "Market Cap ($B)": 0, "Avg Vol (M)": 0, "Beta": 1,
                         "Swing Score": 0, "Vol Score": 0, "Mom Score": 0, "Liq Score": 0,
                         "52wk High": 0, "52wk Low": 0})

    df = pd.DataFrame(rows).sort_values("Swing Score", ascending=False).reset_index(drop=True)
    df.insert(0, "Rank", range(1, len(df) + 1))
    return df


def fetch_buy_signals():
    """Fetch 6 months of daily history and compute technical buy signals."""
    log.info("Computing buy signals for %s", TICKERS)
    signals = []

    for ticker in TICKERS:
        try:
            tk = yf.Ticker(ticker)
            hist = tk.history(period="6mo")
            if hist.empty or len(hist) < 50:
                log.warning("Insufficient history for %s", ticker)
                continue

            close = hist["Close"]
            volume = hist["Volume"]
            info = tk.info or {}
            current = float(close.iloc[-1])
            mkt_cap = info.get("marketCap", 0) / 1e9

            if mkt_cap < 1.0:
                continue

            # --- Technical Indicators ---
            rsi = calc_rsi(close)
            rsi_now = round(float(rsi.iloc[-1]), 1) if not pd.isna(rsi.iloc[-1]) else 50

            macd_line, signal_line, macd_hist = calc_macd(close)
            macd_now = round(float(macd_line.iloc[-1]), 2) if not pd.isna(macd_line.iloc[-1]) else 0
            signal_now = round(float(signal_line.iloc[-1]), 2) if not pd.isna(signal_line.iloc[-1]) else 0
            macd_hist_now = round(float(macd_hist.iloc[-1]), 2) if not pd.isna(macd_hist.iloc[-1]) else 0
            macd_hist_prev = round(float(macd_hist.iloc[-2]), 2) if len(macd_hist) > 1 and not pd.isna(macd_hist.iloc[-2]) else 0
            macd_crossover = "BULLISH" if macd_hist_now > 0 and macd_hist_prev <= 0 else (
                "BEARISH" if macd_hist_now < 0 and macd_hist_prev >= 0 else "NONE")

            bb_upper, bb_mid, bb_lower = calc_bollinger(close)
            bb_upper_now = float(bb_upper.iloc[-1]) if not pd.isna(bb_upper.iloc[-1]) else current
            bb_lower_now = float(bb_lower.iloc[-1]) if not pd.isna(bb_lower.iloc[-1]) else current
            bb_mid_now = float(bb_mid.iloc[-1]) if not pd.isna(bb_mid.iloc[-1]) else current
            bb_width = bb_upper_now - bb_lower_now
            bb_position = ((current - bb_lower_now) / bb_width * 100) if bb_width > 0 else 50
            bb_position = round(min(max(bb_position, 0), 100), 1)

            sma_50 = calc_sma(close, 50)
            sma_200 = calc_sma(close, 200)
            sma50_now = round(float(sma_50.iloc[-1]), 2) if not pd.isna(sma_50.iloc[-1]) else current
            sma200_now = round(float(sma_200.iloc[-1]), 2) if len(sma_200) > 0 and not pd.isna(sma_200.iloc[-1]) else current

            above_50 = current > sma50_now
            above_200 = current > sma200_now
            golden_cross = sma50_now > sma200_now

            avg_vol_20 = float(volume.rolling(20).mean().iloc[-1]) if len(volume) >= 20 else float(volume.mean())
            vol_today = float(volume.iloc[-1])
            vol_ratio = round(vol_today / avg_vol_20, 2) if avg_vol_20 > 0 else 1.0
            vol_surge = vol_ratio > 1.5

            fifty_two_high = float(close.max())
            fifty_two_low = float(close.min())
            dist_from_high = round(((current - fifty_two_high) / fifty_two_high) * 100, 1)
            dist_from_low = round(((current - fifty_two_low) / fifty_two_low) * 100, 1)

            # --- BUY SCORE CALCULATION (0-100) ---
            buy_score = 0

            # RSI component (0-25): oversold = higher score
            if rsi_now < 30:
                buy_score += 25
            elif rsi_now < 40:
                buy_score += 20
            elif rsi_now < 50:
                buy_score += 15
            elif rsi_now < 60:
                buy_score += 10
            elif rsi_now < 70:
                buy_score += 5
            # Overbought (>70) adds 0

            # MACD component (0-20)
            if macd_crossover == "BULLISH":
                buy_score += 20
            elif macd_hist_now > 0 and macd_hist_now > macd_hist_prev:
                buy_score += 15
            elif macd_hist_now > 0:
                buy_score += 10
            elif macd_hist_now < 0 and macd_hist_now > macd_hist_prev:
                buy_score += 5

            # Bollinger Band component (0-20): near lower band = buy
            if bb_position < 10:
                buy_score += 20
            elif bb_position < 25:
                buy_score += 15
            elif bb_position < 40:
                buy_score += 10
            elif bb_position < 60:
                buy_score += 5

            # Moving Average component (0-20)
            if golden_cross and above_50 and above_200:
                buy_score += 20
            elif above_50 and above_200:
                buy_score += 15
            elif above_50:
                buy_score += 10
            elif above_200:
                buy_score += 5

            # Volume surge component (0-15)
            if vol_surge and current > float(close.iloc[-2]):
                buy_score += 15
            elif vol_surge:
                buy_score += 8
            elif vol_ratio > 1.2 and current > float(close.iloc[-2]):
                buy_score += 5

            buy_score = min(buy_score, 100)

            # Rating
            if buy_score >= 70:
                rating = "BUY"
            elif buy_score >= 45:
                rating = "WATCH"
            else:
                rating = "AVOID"

            # Key reason
            reasons = []
            if rsi_now < 35: reasons.append(f"RSI oversold ({rsi_now})")
            if macd_crossover == "BULLISH": reasons.append("MACD bullish crossover")
            if bb_position < 20: reasons.append(f"Near Bollinger low ({bb_position}%)")
            if golden_cross: reasons.append("Golden cross (50>200 SMA)")
            if vol_surge and current > float(close.iloc[-2]): reasons.append(f"Volume surge {vol_ratio}x on up day")
            if dist_from_high > -5: reasons.append(f"Near 52wk high ({dist_from_high}%)")
            if rsi_now > 70: reasons.append(f"RSI overbought ({rsi_now})")
            if macd_crossover == "BEARISH": reasons.append("MACD bearish crossover")
            if not reasons:
                reasons.append("No strong signal")

            signals.append({
                "Ticker": ticker,
                "Company": info.get("shortName", ticker),
                "Sector": SECTORS.get(ticker, "N/A"),
                "Price": round(current, 2),
                "Mkt Cap ($B)": round(mkt_cap, 1),
                "RSI (14)": rsi_now,
                "MACD": macd_now,
                "MACD Signal": signal_now,
                "MACD Cross": macd_crossover,
                "BB Position %": bb_position,
                "50-SMA": sma50_now,
                "200-SMA": sma200_now,
                "Above 50-SMA": "Yes" if above_50 else "No",
                "Above 200-SMA": "Yes" if above_200 else "No",
                "Golden Cross": "Yes" if golden_cross else "No",
                "Vol Ratio": vol_ratio,
                "Vol Surge": "YES" if vol_surge else "No",
                "Dist from 52wk High": dist_from_high,
                "Dist from 52wk Low": dist_from_low,
                "Buy Score": buy_score,
                "Rating": rating,
                "Key Reasons": " | ".join(reasons[:3]),
            })
        except Exception as e:
            log.warning("Buy signal error %s: %s", ticker, e)

    return sorted(signals, key=lambda x: x["Buy Score"], reverse=True)


def fetch_futures_data():
    log.info("Fetching futures & macro data")
    symbols = {
        "ES=F": "S&P 500 Futures", "NQ=F": "Nasdaq 100 Futures",
        "YM=F": "Dow Futures", "RTY=F": "Russell 2000 Futures",
        "^VIX": "VIX (Fear Index)", "^TNX": "10-Year Treasury Yield",
        "DX-Y.NYB": "US Dollar Index (DXY)",
    }
    results = []
    for sym, name in symbols.items():
        try:
            tk = yf.Ticker(sym)
            info = tk.info or {}
            current = info.get("regularMarketPrice") or info.get("previousClose", 0)
            prev = info.get("regularMarketPreviousClose") or info.get("previousClose", current)
            change = current - prev
            pct = (change / prev * 100) if prev else 0
            if "VIX" in name:
                signal = "DECREASING FEAR" if change < 0 else "INCREASING FEAR"
            elif pct > 0.3: signal = "BULLISH"
            elif pct > 0: signal = "SLIGHTLY BULLISH"
            elif pct > -0.3: signal = "NEUTRAL"
            else: signal = "BEARISH"
            results.append({"name": name, "level": round(current, 2),
                            "change": round(change, 2), "pct": round(pct, 2), "signal": signal})
        except Exception as e:
            log.warning("Futures error %s: %s", sym, e)
            results.append({"name": name, "level": 0, "change": 0, "pct": 0, "signal": "N/A"})
    return results


def fetch_market_news_context():
    return {
        "fed_rate": "3.50% - 3.75%",
        "fed_status": "Rates held at Jan 27-28 meeting. 2 dissents favored cut.",
        "next_fomc": "March 17-18, 2026",
        "market_expects": "93% prob rates held in March; ~65bps cuts priced for 2026",
        "overall_market": "BULL MARKET — S&P 500 near all-time highs",
    }


# ===========================================================================
#  CHART GENERATION
# ===========================================================================

def generate_charts(tickers_list):
    """Generate technical analysis chart images for each ticker. Returns dict of ticker -> png path and base64."""
    log.info("Generating technical charts for %s", tickers_list)
    charts = {}
    charts_dir = os.path.join(CONFIG["OUTPUT_DIR"], "charts")
    os.makedirs(charts_dir, exist_ok=True)

    plt.rcParams.update({
        "figure.facecolor": "#0d1117", "axes.facecolor": "#161b22",
        "axes.edgecolor": "#30363d", "axes.labelcolor": "#e6edf3",
        "text.color": "#e6edf3", "xtick.color": "#8b949e", "ytick.color": "#8b949e",
        "grid.color": "#21262d", "grid.alpha": 0.5,
    })

    for ticker in tickers_list:
        try:
            tk = yf.Ticker(ticker)
            hist = tk.history(period="6mo")
            if hist.empty or len(hist) < 30:
                continue

            close = hist["Close"]
            volume = hist["Volume"]
            dates = hist.index

            # Indicators
            sma50 = calc_sma(close, 50)
            sma200 = calc_sma(close, 200)
            bb_upper, bb_mid, bb_lower = calc_bollinger(close)
            rsi = calc_rsi(close)
            macd_line, signal_line, macd_histogram = calc_macd(close)

            # Create figure with 4 subplots
            fig = plt.figure(figsize=(14, 10))
            gs = GridSpec(4, 1, height_ratios=[3, 1, 1, 1], hspace=0.08, top=0.93, bottom=0.06, left=0.08, right=0.96)

            # --- Price + MAs + Bollinger ---
            ax1 = fig.add_subplot(gs[0])
            ax1.plot(dates, close, color="#58a6ff", linewidth=1.5, label="Price")
            ax1.plot(dates, sma50, color="#f0883e", linewidth=1, alpha=0.8, label="50-SMA")
            ax1.plot(dates, sma200, color="#d2a8ff", linewidth=1, alpha=0.8, label="200-SMA")
            ax1.fill_between(dates, bb_upper, bb_lower, alpha=0.1, color="#58a6ff", label="Bollinger Bands")
            ax1.plot(dates, bb_upper, color="#58a6ff", linewidth=0.5, alpha=0.4)
            ax1.plot(dates, bb_lower, color="#58a6ff", linewidth=0.5, alpha=0.4)

            info = tk.info or {}
            rating = "N/A"
            for s in latest.get("buy_signals", []):
                if s["Ticker"] == ticker:
                    rating = s["Rating"]
                    break
            r_color = "#3fb950" if rating == "BUY" else ("#d29922" if rating == "WATCH" else "#f85149")
            ax1.set_title(f"{ticker} — {info.get('shortName', ticker)}  |  Rating: {rating}",
                          fontsize=16, fontweight="bold", color=r_color, pad=10)
            ax1.legend(loc="upper left", fontsize=8, facecolor="#161b22", edgecolor="#30363d", labelcolor="#e6edf3")
            ax1.set_ylabel("Price ($)", fontsize=10)
            ax1.tick_params(labelbottom=False)
            ax1.grid(True)

            # --- Volume ---
            ax2 = fig.add_subplot(gs[1], sharex=ax1)
            colors = ["#3fb950" if close.iloc[i] >= close.iloc[i-1] else "#f85149" for i in range(1, len(close))]
            colors.insert(0, "#3fb950")
            ax2.bar(dates, volume, color=colors, alpha=0.7, width=0.8)
            avg_vol = volume.rolling(20).mean()
            ax2.plot(dates, avg_vol, color="#d29922", linewidth=1, alpha=0.8, label="20-day Avg")
            ax2.set_ylabel("Volume", fontsize=9)
            ax2.tick_params(labelbottom=False)
            ax2.legend(loc="upper left", fontsize=7, facecolor="#161b22", edgecolor="#30363d", labelcolor="#e6edf3")
            ax2.grid(True)

            # --- RSI ---
            ax3 = fig.add_subplot(gs[2], sharex=ax1)
            ax3.plot(dates, rsi, color="#58a6ff", linewidth=1.2)
            ax3.axhline(y=70, color="#f85149", linewidth=0.8, linestyle="--", alpha=0.7)
            ax3.axhline(y=30, color="#3fb950", linewidth=0.8, linestyle="--", alpha=0.7)
            ax3.fill_between(dates, rsi, 70, where=(rsi > 70), alpha=0.2, color="#f85149")
            ax3.fill_between(dates, rsi, 30, where=(rsi < 30), alpha=0.2, color="#3fb950")
            ax3.set_ylabel("RSI (14)", fontsize=9)
            ax3.set_ylim(10, 90)
            ax3.tick_params(labelbottom=False)
            ax3.grid(True)

            # --- MACD ---
            ax4 = fig.add_subplot(gs[3], sharex=ax1)
            ax4.plot(dates, macd_line, color="#58a6ff", linewidth=1, label="MACD")
            ax4.plot(dates, signal_line, color="#f0883e", linewidth=1, label="Signal")
            hist_colors = ["#3fb950" if v >= 0 else "#f85149" for v in macd_histogram]
            ax4.bar(dates, macd_histogram, color=hist_colors, alpha=0.5, width=0.8)
            ax4.axhline(y=0, color="#30363d", linewidth=0.5)
            ax4.set_ylabel("MACD", fontsize=9)
            ax4.legend(loc="upper left", fontsize=7, facecolor="#161b22", edgecolor="#30363d", labelcolor="#e6edf3")
            ax4.xaxis.set_major_formatter(mdates.DateFormatter("%b %d"))
            ax4.xaxis.set_major_locator(mdates.MonthLocator())
            plt.setp(ax4.xaxis.get_majorticklabels(), rotation=45, fontsize=8)
            ax4.grid(True)

            # Save
            png_path = os.path.join(charts_dir, f"{ticker}_chart.png")
            fig.savefig(png_path, dpi=130, facecolor="#0d1117")
            plt.close(fig)

            # Base64 for web
            with open(png_path, "rb") as f:
                b64 = base64.b64encode(f.read()).decode("utf-8")

            charts[ticker] = {"path": png_path, "base64": b64}
            log.info("Chart generated: %s", ticker)
        except Exception as e:
            log.warning("Chart error %s: %s", ticker, e)

    return charts


# ===========================================================================
#  SPREADSHEET GENERATION
# ===========================================================================

def build_spreadsheet(df, futures, macro, buy_signals, charts=None):
    wb = Workbook()
    now = datetime.now(CST)
    timestamp = now.strftime("%Y%m%d_%H%M")

    # Styles
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    dfont = Font(name="Arial", size=10)
    bfont = Font(name="Arial", bold=True, size=10)
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    orange_fill = PatternFill("solid", fgColor="FFF2CC")
    sub_fill = PatternFill("solid", fgColor="D6E4F0")
    green_font = Font(name="Arial", size=10, color="006100")
    red_font = Font(name="Arial", size=10, color="C00000")
    bdr = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    def sc(ws, r, c, val, font=dfont, fill=None, align="center", fmt=None):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = font; cell.border = bdr
        cell.alignment = Alignment(horizontal=align, wrap_text=True)
        if fill: cell.fill = fill
        if fmt: cell.number_format = fmt
        return cell

    # ============================
    # SHEET 1: MARKET OVERVIEW
    # ============================
    ws1 = wb.active
    ws1.title = "Market Overview"
    ws1.merge_cells("A1:G1")
    ws1["A1"] = "SWING TRADE MARKET INTELLIGENCE DASHBOARD"
    ws1["A1"].font = Font(name="Arial", bold=True, size=16, color="1F4E79")
    ws1["A1"].alignment = Alignment(horizontal="center")
    ws1.merge_cells("A2:G2")
    ws1["A2"] = f"Generated: {now.strftime('%B %d, %Y at %I:%M %p %Z')}"
    ws1["A2"].font = Font(name="Arial", italic=True, size=10, color="666666")
    ws1["A2"].alignment = Alignment(horizontal="center")

    # Futures
    r = 4
    ws1.merge_cells(f"A{r}:G{r}")
    c = ws1[f"A{r}"]; c.value = "PRE-MARKET / CURRENT FUTURES & MACRO INDICATORS"
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2E75B6"); c.alignment = Alignment(horizontal="center")
    r = 5
    for i, h in enumerate(["Index / Future", "Level", "Change", "% Change", "Signal", "", ""], 1):
        sc(ws1, r, i, h, font=hdr_font, fill=hdr_fill)
    for idx, fut in enumerate(futures):
        row = 6 + idx
        sc(ws1, row, 1, fut["name"], align="left")
        sc(ws1, row, 2, fut["level"], fmt="#,##0.00")
        ch = fut["change"]
        sc(ws1, row, 3, f"{'+' if ch>=0 else ''}{ch}", font=green_font if ch>=0 else red_font)
        sc(ws1, row, 4, f"{'+' if fut['pct']>=0 else ''}{fut['pct']}%", font=green_font if fut["pct"]>=0 else red_font)
        sig = fut["signal"]
        sig_fill = green_fill if "BULL" in sig or "DECREASING" in sig else (red_fill if "BEAR" in sig or "INCREASING" in sig else sub_fill)
        sc(ws1, row, 5, sig, font=bfont, fill=sig_fill)

    bull_count = sum(1 for f in futures if "BULL" in f["signal"] or "DECREASING" in f["signal"])
    verdict = "BULLISH" if bull_count >= 4 else ("MIXED" if bull_count >= 2 else "BEARISH")
    vr = 6 + len(futures)
    ws1.merge_cells(f"A{vr}:G{vr}")
    ws1[f"A{vr}"] = f"FUTURES VERDICT: {verdict} ({bull_count}/{len(futures)} indicators positive)"
    ws1[f"A{vr}"].font = Font(name="Arial", bold=True, size=10, color="006100" if verdict=="BULLISH" else "C00000" if verdict=="BEARISH" else "BF8F00")
    ws1[f"A{vr}"].fill = green_fill if verdict=="BULLISH" else (red_fill if verdict=="BEARISH" else sub_fill)

    # Fed/macro
    r = vr + 2
    ws1.merge_cells(f"A{r}:G{r}")
    c = ws1[f"A{r}"]; c.value = "OVERALL MARKET & FEDERAL RESERVE STATUS"
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="C00000"); c.alignment = Alignment(horizontal="center")
    r += 1
    for i, h in enumerate(["Item", "Value", "Status", "", "", "", ""], 1):
        sc(ws1, r, i, h, font=hdr_font, fill=PatternFill("solid", fgColor="8B0000"))
    macro_rows = [
        ["Overall Market", macro["overall_market"], "BULL"],
        ["Fed Funds Rate", macro["fed_rate"], "HELD"],
        ["Last FOMC Decision", macro["fed_status"], "DIVIDED"],
        ["Next FOMC Meeting", macro["next_fomc"], "UPCOMING"],
        ["Rate Cut Expectations", macro["market_expects"], "DOVISH LEAN"],
    ]
    for idx, mr in enumerate(macro_rows):
        row = r + 1 + idx
        sc(ws1, row, 1, mr[0], align="left")
        sc(ws1, row, 2, mr[1], align="left")
        fill = green_fill if mr[2] in ("BULL","DOVISH LEAN") else (red_fill if mr[2]=="BEAR" else sub_fill)
        sc(ws1, row, 3, mr[2], font=bfont, fill=fill)

    for col, w in enumerate([35, 28, 25, 20, 35, 45, 5], 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ============================
    # SHEET 2: TOP 10 SWING TRADES
    # ============================
    ws2 = wb.create_sheet("Top 10 Swing Trades")
    ws2.merge_cells("A1:P1")
    ws2["A1"] = "TOP 10 NASDAQ SWING TRADE OPPORTUNITIES (Market Cap > $1B)"
    ws2["A1"].font = Font(name="Arial", bold=True, size=16, color="1F4E79")
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.merge_cells("A2:P2")
    ws2["A2"] = f"Updated: {now.strftime('%B %d, %Y at %I:%M %p %Z')} | Top 3 highlighted in yellow"
    ws2["A2"].font = Font(name="Arial", italic=True, size=10, color="666666")
    ws2["A2"].alignment = Alignment(horizontal="center")

    cols = ["Rank","Ticker","Company","Sector","Current Price","Prev Close",
            "3-Mo Ago Price","Daily % Chg","3-Month % Chg","Market Cap ($B)",
            "Avg Vol (M)","Swing Score","Vol Score","Mom Score","Liq Score","52wk Range"]
    for i, h in enumerate(cols, 1):
        sc(ws2, 3, i, h, font=hdr_font, fill=hdr_fill)
    for idx, (_, rd) in enumerate(df.iterrows()):
        r = 4 + idx
        top3 = rd["Rank"] <= 3
        rf = yellow_fill if top3 else None
        sc(ws2, r, 1, int(rd["Rank"]), fill=rf)
        sc(ws2, r, 2, rd["Ticker"], font=bfont, fill=rf)
        sc(ws2, r, 3, rd["Company"], align="left", fill=rf)
        sc(ws2, r, 4, rd["Sector"], fill=rf)
        sc(ws2, r, 5, rd["Current Price"], fmt="$#,##0.00", fill=rf)
        sc(ws2, r, 6, rd["Prev Close"], fmt="$#,##0.00", fill=rf)
        sc(ws2, r, 7, rd["3-Mo Ago Price"], fmt="$#,##0.00", fill=rf)
        dp = rd["Daily % Chg"]
        sc(ws2, r, 8, dp, font=Font(name="Arial", bold=True, size=10, color="006100" if dp>=0 else "C00000"),
           fill=rf if top3 else (green_fill if dp>=0 else red_fill), fmt="0.00%")
        mp = rd["3-Month % Chg"]
        sc(ws2, r, 9, mp, font=Font(name="Arial", bold=True, size=10, color="006100" if mp>=0 else "C00000"),
           fill=rf if top3 else (green_fill if mp>=0 else red_fill), fmt="0.00%")
        sc(ws2, r, 10, rd["Market Cap ($B)"], fmt="#,##0.0", fill=rf)
        sc(ws2, r, 11, rd["Avg Vol (M)"], fmt="#,##0.0", fill=rf)
        ss = rd["Swing Score"]
        sc(ws2, r, 12, ss, font=bfont, fill=rf if top3 else (green_fill if ss>=80 else sub_fill))
        sc(ws2, r, 13, rd["Vol Score"], fill=rf)
        sc(ws2, r, 14, rd["Mom Score"], fill=rf)
        sc(ws2, r, 15, rd["Liq Score"], fill=rf)
        sc(ws2, r, 16, f"${rd['52wk Low']:.2f} - ${rd['52wk High']:.2f}", fill=rf)
    for col, w in enumerate([6,8,28,18,13,13,13,12,13,14,12,12,10,10,10,22], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ============================
    # SHEET 3: BUY SIGNALS  (NEW)
    # ============================
    ws3 = wb.create_sheet("Buy Signals")
    ws3.merge_cells("A1:V1")
    ws3["A1"] = "TECHNICAL BUY SIGNAL ANALYSIS — NASDAQ Stocks (Mkt Cap > $1B)"
    ws3["A1"].font = Font(name="Arial", bold=True, size=16, color="1F4E79")
    ws3["A1"].alignment = Alignment(horizontal="center")
    ws3.merge_cells("A2:V2")
    ws3["A2"] = f"Updated: {now.strftime('%B %d, %Y at %I:%M %p %Z')} | BUY = Score 70+ | WATCH = 45-69 | AVOID = Below 45"
    ws3["A2"].font = Font(name="Arial", italic=True, size=10, color="666666")
    ws3["A2"].alignment = Alignment(horizontal="center")

    buy_cols = ["Ticker", "Company", "Sector", "Price", "Mkt Cap ($B)",
                "Buy Score", "Rating",
                "RSI (14)", "MACD", "MACD Signal", "MACD Cross",
                "BB Position %", "50-SMA", "200-SMA",
                "Above 50-SMA", "Above 200-SMA", "Golden Cross",
                "Vol Ratio", "Vol Surge",
                "From 52wk High", "From 52wk Low",
                "Key Reasons"]
    for i, h in enumerate(buy_cols, 1):
        sc(ws3, 3, i, h, font=hdr_font, fill=hdr_fill)

    for idx, sig in enumerate(buy_signals):
        r = 4 + idx
        rating = sig["Rating"]
        if rating == "BUY":
            row_fill = green_fill
        elif rating == "WATCH":
            row_fill = orange_fill
        else:
            row_fill = red_fill

        sc(ws3, r, 1, sig["Ticker"], font=bfont, fill=row_fill)
        sc(ws3, r, 2, sig["Company"], align="left", fill=row_fill)
        sc(ws3, r, 3, sig["Sector"], fill=row_fill)
        sc(ws3, r, 4, sig["Price"], fmt="$#,##0.00", fill=row_fill)
        sc(ws3, r, 5, sig["Mkt Cap ($B)"], fmt="#,##0.0", fill=row_fill)

        # Buy Score - bold colored
        bs = sig["Buy Score"]
        bs_fill = green_fill if bs >= 70 else (orange_fill if bs >= 45 else red_fill)
        sc(ws3, r, 6, bs, font=Font(name="Arial", bold=True, size=12), fill=bs_fill)

        # Rating - big bold
        r_color = "006100" if rating == "BUY" else ("BF8F00" if rating == "WATCH" else "C00000")
        sc(ws3, r, 7, rating, font=Font(name="Arial", bold=True, size=12, color=r_color), fill=row_fill)

        # RSI
        rsi_v = sig["RSI (14)"]
        rsi_fill = green_fill if rsi_v < 30 else (red_fill if rsi_v > 70 else row_fill)
        sc(ws3, r, 8, rsi_v, fmt="0.0", fill=rsi_fill)

        sc(ws3, r, 9, sig["MACD"], fmt="0.00", fill=row_fill)
        sc(ws3, r, 10, sig["MACD Signal"], fmt="0.00", fill=row_fill)

        mc = sig["MACD Cross"]
        mc_fill = green_fill if mc == "BULLISH" else (red_fill if mc == "BEARISH" else row_fill)
        sc(ws3, r, 11, mc, font=bfont, fill=mc_fill)

        # BB Position
        bb = sig["BB Position %"]
        bb_fill = green_fill if bb < 20 else (red_fill if bb > 80 else row_fill)
        sc(ws3, r, 12, bb, fmt="0.0", fill=bb_fill)

        sc(ws3, r, 13, sig["50-SMA"], fmt="$#,##0.00", fill=row_fill)
        sc(ws3, r, 14, sig["200-SMA"], fmt="$#,##0.00", fill=row_fill)

        for col_idx, key in enumerate(["Above 50-SMA", "Above 200-SMA", "Golden Cross"], 15):
            val = sig[key]
            f = green_fill if val == "Yes" else row_fill
            sc(ws3, r, col_idx, val, font=bfont, fill=f)

        sc(ws3, r, 18, sig["Vol Ratio"], fmt="0.00", fill=row_fill)
        vs = sig["Vol Surge"]
        sc(ws3, r, 19, vs, font=bfont, fill=green_fill if vs == "YES" else row_fill)

        dh = sig["Dist from 52wk High"]
        sc(ws3, r, 20, dh, fmt="0.0%", fill=row_fill,
           font=green_font if dh > -5 else red_font)
        dl = sig["Dist from 52wk Low"]
        sc(ws3, r, 21, dl, fmt="0.0%", fill=row_fill, font=green_font)
        sc(ws3, r, 22, sig["Key Reasons"], align="left", fill=row_fill)

    # Legend
    lr = 4 + len(buy_signals) + 2
    ws3.merge_cells(f"A{lr}:V{lr}")
    ws3[f"A{lr}"] = "BUY SCORE METHODOLOGY (0-100)"
    ws3[f"A{lr}"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws3[f"A{lr}"].fill = PatternFill("solid", fgColor="2E75B6")
    ws3[f"A{lr}"].alignment = Alignment(horizontal="center")

    legend = [
        ["Component", "Max Points", "What It Measures", "Buy Signal"],
        ["RSI (14-day)", "25 pts", "Overbought/oversold momentum", "Below 30 = oversold = strong buy signal"],
        ["MACD Crossover", "20 pts", "Trend change detection", "MACD crossing above signal line = bullish"],
        ["Bollinger Band Position", "20 pts", "Price vs statistical range", "Near lower band (<20%) = potential bounce"],
        ["Moving Averages (50/200)", "20 pts", "Trend direction & support", "Golden cross + above both SMAs = strong uptrend"],
        ["Volume Surge", "15 pts", "Institutional buying pressure", "Volume >1.5x average on up day = accumulation"],
        ["", "", "", ""],
        ["RATINGS:", "", "BUY (70-100) = Strong technicals, consider entry", "WATCH (45-69) = Mixed signals, wait for confirmation"],
        ["", "", "AVOID (0-44) = Weak technicals, stay away for now", ""],
    ]
    for i, row in enumerate(legend):
        for j, val in enumerate(row):
            cell = sc(ws3, lr+1+i, j+1, val, font=bfont if i==0 else dfont)
            if i == 0: cell.fill = sub_fill

    for col, w in enumerate([8,28,18,12,12,11,10,9,10,11,13,12,12,12,12,13,12,10,10,13,13,40], 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    # ============================
    # SHEET 4: TECHNICAL CHARTS
    # ============================
    if charts:
        ws_charts = wb.create_sheet("Technical Charts")
        ws_charts.merge_cells("A1:N1")
        ws_charts["A1"] = "TECHNICAL ANALYSIS CHARTS — 6-Month Price + Indicators"
        ws_charts["A1"].font = Font(name="Arial", bold=True, size=16, color="1F4E79")
        ws_charts["A1"].alignment = Alignment(horizontal="center")
        ws_charts.merge_cells("A2:N2")
        ws_charts["A2"] = f"Updated: {now.strftime('%B %d, %Y at %I:%M %p %Z')} | Each chart: Price + Bollinger + 50/200 SMA + Volume + RSI + MACD"
        ws_charts["A2"].font = Font(name="Arial", italic=True, size=10, color="666666")
        ws_charts["A2"].alignment = Alignment(horizontal="center")

        chart_row = 4
        for i, ticker in enumerate(TICKERS):
            if ticker in charts:
                # Add label
                ws_charts.merge_cells(f"A{chart_row}:N{chart_row}")
                label_cell = ws_charts[f"A{chart_row}"]
                rating = "N/A"
                for s in buy_signals:
                    if s["Ticker"] == ticker:
                        rating = s["Rating"]
                        break
                label_cell.value = f"{ticker} — Rating: {rating}"
                r_color = "006100" if rating == "BUY" else ("BF8F00" if rating == "WATCH" else "C00000")
                label_cell.font = Font(name="Arial", bold=True, size=14, color=r_color)
                label_cell.alignment = Alignment(horizontal="center")

                chart_row += 1
                try:
                    img = XlImage(charts[ticker]["path"])
                    img.width = 980
                    img.height = 700
                    ws_charts.add_image(img, f"A{chart_row}")
                except Exception as e:
                    log.warning("Could not embed chart for %s: %s", ticker, e)
                    ws_charts[f"A{chart_row}"] = f"Chart unavailable for {ticker}"

                chart_row += 38  # space for the image (~700px at default row height)

        ws_charts.column_dimensions['A'].width = 12
        for c in range(2, 15):
            ws_charts.column_dimensions[get_column_letter(c)].width = 12

    # ============================
    # SHEET 5: TRADING NOTES
    # ============================
    ws4 = wb.create_sheet("Trading Notes")
    ws4.merge_cells("A1:D1")
    ws4["A1"] = "TRADING NOTES & RISK FACTORS"
    ws4["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    notes = [
        ["", ""],
        ["HOW TO USE THIS DASHBOARD", ""],
        ["1.", "Check the Market Overview tab first — futures and Fed status set the day's tone"],
        ["2.", "Review Top 10 Swing Trades — sorted by Swing Score (volatility + momentum + volume)"],
        ["3.", "Check the Buy Signals tab — stocks with BUY rating have the strongest technical setup"],
        ["4.", "NEW: Check Technical Charts tab for visual confirmation of signals"],
        ["5.", "Cross-reference: HIGH Swing Score + BUY rating + chart confirmation = best opportunity"],
        ["6.", "Always check if earnings are within 7 days — volatility spikes around reports"],
        ["", ""],
        ["RISK MANAGEMENT RULES", ""],
        ["•", "Never risk more than 1-2% of your account on a single swing trade"],
        ["•", "Always set a stop-loss before entering a trade"],
        ["•", "Take partial profits at 1:1 risk/reward, let remainder run"],
        ["•", "Avoid entering new positions right before major catalysts (earnings, Fed)"],
        ["•", "Reduce position sizes when VIX is above 25"],
        ["", ""],
        ["DISCLAIMER", ""],
        ["", "This spreadsheet is for informational purposes only. Not financial advice."],
        ["", "Swing trading involves significant risk. Past performance doesn't guarantee future results."],
    ]
    for i, row in enumerate(notes):
        r = 2 + i
        for j, val in enumerate(row):
            ws4.cell(row=r, column=j+1, value=val).font = dfont
        if any(k in str(row[0]) for k in ["HOW TO USE", "RISK MANAGEMENT", "DISCLAIMER"]):
            ws4.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=12, color="C00000")
    ws4.column_dimensions['A'].width = 5
    ws4.column_dimensions['B'].width = 80

    # Save
    os.makedirs(CONFIG["OUTPUT_DIR"], exist_ok=True)
    fname = f"swing_dashboard_{timestamp}.xlsx"
    fpath = os.path.join(CONFIG["OUTPUT_DIR"], fname)
    wb.save(fpath)
    wb.save(os.path.join(CONFIG["OUTPUT_DIR"], "latest.xlsx"))
    log.info("Spreadsheet saved: %s", fpath)
    return fpath


# ===========================================================================
#  EMAIL
# ===========================================================================

def send_email(filepath, skip=False):
    if skip:
        log.info("Email skipped (--no-email)")
        return
    user = CONFIG["SMTP_USER"]
    pwd = CONFIG["SMTP_PASSWORD"]
    if not user or not pwd:
        log.warning("SMTP credentials not set — skipping email. Set SMTP_USER and SMTP_PASSWORD env vars.")
        return

    msg = MIMEMultipart()
    msg["From"] = user
    msg["To"] = CONFIG["EMAIL_TO"]
    now_str = datetime.now(CST).strftime("%I:%M %p CST on %b %d, %Y")
    msg["Subject"] = f"Swing Trade Dashboard — {now_str}"

    # Build summary from buy signals
    buy_list = [s for s in latest.get("buy_signals", []) if s["Rating"] == "BUY"]
    watch_list = [s for s in latest.get("buy_signals", []) if s["Rating"] == "WATCH"]

    body = f"Swing Trade Dashboard refreshed at {now_str}.\n\n"
    if buy_list:
        body += "BUY SIGNALS:\n"
        for s in buy_list:
            body += f"  {s['Ticker']} — Score {s['Buy Score']} — {s['Key Reasons']}\n"
        body += "\n"
    if watch_list:
        body += "WATCH LIST:\n"
        for s in watch_list:
            body += f"  {s['Ticker']} — Score {s['Buy Score']} — {s['Key Reasons']}\n"
        body += "\n"
    body += "See attached spreadsheet for full analysis.\n\n— Swing Trade Bot"
    msg.attach(MIMEText(body, "plain"))

    with open(filepath, "rb") as f:
        part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(filepath)}")
        msg.attach(part)

    try:
        with smtplib.SMTP(CONFIG["SMTP_SERVER"], CONFIG["SMTP_PORT"]) as server:
            server.ehlo()
            server.starttls()
            server.login(user, pwd)
            server.sendmail(user, CONFIG["EMAIL_TO"], msg.as_string())
        log.info("Email sent to %s", CONFIG["EMAIL_TO"])
    except Exception as e:
        log.error("Email failed: %s", e)


# ===========================================================================
#  MAIN JOB
# ===========================================================================

def run_job(skip_email=False):
    log.info("=" * 60)
    log.info("RUNNING SWING TRADE DASHBOARD JOB v2.0")
    log.info("=" * 60)
    try:
        df = fetch_stock_data()
        futures = fetch_futures_data()
        macro = fetch_market_news_context()
        buy_signals = fetch_buy_signals()

        latest["buy_signals"] = buy_signals
        charts = generate_charts(TICKERS)

        filepath = build_spreadsheet(df, futures, macro, buy_signals, charts)

        latest["file"] = filepath
        latest["generated_at"] = datetime.now(CST).isoformat()
        latest["chart_images"] = {t: c["base64"] for t, c in charts.items()}
        latest["summary"] = {
            "top3": df.head(3)[["Ticker","Swing Score","Daily % Chg","Current Price"]].to_dict("records"),
            "futures_verdict": "BULLISH" if sum(1 for f in futures if "BULL" in f["signal"] or "DECREASING" in f["signal"]) >= 4 else "MIXED",
            "futures": futures,
        }
        latest["error"] = None

        send_email(filepath, skip=skip_email)
        log.info("Job complete: %s", filepath)
    except Exception as e:
        log.error("Job failed: %s", e, exc_info=True)
        latest["error"] = str(e)


# ===========================================================================
#  WEB DASHBOARD
# ===========================================================================

app = Flask(__name__)

DASHBOARD_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Swing Trade Dashboard v2.0</title>
<style>
  *{box-sizing:border-box;margin:0;padding:0}
  body{font-family:'Segoe UI',Arial,sans-serif;background:#0d1117;color:#e6edf3}
  .header{background:linear-gradient(135deg,#1f4e79,#2e75b6);padding:24px;text-align:center}
  .header h1{font-size:28px;margin-bottom:4px}.header p{opacity:.8;font-size:14px}
  .container{max-width:1400px;margin:20px auto;padding:0 16px}
  .card{background:#161b22;border:1px solid #30363d;border-radius:8px;padding:20px;margin-bottom:16px}
  .card h2{color:#58a6ff;margin-bottom:12px;font-size:18px}
  table{width:100%;border-collapse:collapse;font-size:14px}
  th{background:#1f4e79;color:white;padding:10px 8px;text-align:center}
  td{padding:8px;text-align:center;border-bottom:1px solid #30363d}
  .pos{color:#3fb950;font-weight:bold}.neg{color:#f85149;font-weight:bold}
  .top3{background:rgba(255,255,0,.12)}
  .buy-row{background:rgba(63,185,80,.1)}.watch-row{background:rgba(210,153,34,.1)}.avoid-row{background:rgba(248,81,73,.1)}
  .badge{display:inline-block;padding:3px 10px;border-radius:12px;font-size:12px;font-weight:bold}
  .bull{background:#0d4429;color:#3fb950}.bear{background:#4a1520;color:#f85149}
  .neutral{background:#3d2e00;color:#d29922}
  .buy-badge{background:#0d4429;color:#3fb950;font-size:14px;padding:4px 14px}
  .watch-badge{background:#3d2e00;color:#d29922;font-size:14px;padding:4px 14px}
  .avoid-badge{background:#4a1520;color:#f85149;font-size:14px;padding:4px 14px}
  .btn{display:inline-block;background:#2e75b6;color:white;padding:10px 24px;border-radius:6px;text-decoration:none;font-weight:bold;margin:4px}
  .btn:hover{background:#1f4e79}
  .actions{text-align:center;margin:16px 0}
  .refresh-info{text-align:center;padding:10px;color:#8b949e;font-size:13px}
  #status{text-align:center;padding:8px;color:#d29922}
  .tabs{display:flex;gap:4px;margin-bottom:16px}.tab{padding:10px 20px;background:#21262d;color:#8b949e;border:1px solid #30363d;border-radius:6px 6px 0 0;cursor:pointer;font-weight:bold}
  .tab.active{background:#161b22;color:#58a6ff;border-bottom-color:#161b22}
  .tab-content{display:none}.tab-content.active{display:block}
</style>
</head>
<body>
<div class="header">
  <h1>Swing Trade Dashboard v2.0</h1>
  <p>Live NASDAQ Intelligence + Technical Buy Signals | Auto-refreshes 4x daily (7am, 9:30am, 12pm, 2:45pm CST)</p>
</div>
<div class="container">
  <div class="actions">
    <a class="btn" href="/download">Download .xlsx</a>
    <a class="btn" href="#" id="refreshBtn">Refresh Now</a>
  </div>
  <div id="status"></div>
  <div class="tabs">
    <div class="tab active" onclick="switchTab('futures')">Futures & Macro</div>
    <div class="tab" onclick="switchTab('swing')">Top 10 Swing Trades</div>
    <div class="tab" onclick="switchTab('buy')">Buy Signals</div>
    <div class="tab" onclick="switchTab('charts')">Technical Charts</div>
  </div>
  <div id="futures" class="tab-content active"></div>
  <div id="swing" class="tab-content"></div>
  <div id="buy" class="tab-content"></div>
  <div id="charts" class="tab-content"></div>
  <div class="refresh-info" id="timestamp"></div>
</div>
<script>
function switchTab(id){
  document.querySelectorAll('.tab-content').forEach(t=>t.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  event.target.classList.add('active');
}
async function loadDashboard(){
  try{
    const res=await fetch('/api/data');
    const d=await res.json();
    if(d.error){document.getElementById('futures').innerHTML=`<div class="card"><p>Error: ${d.error}. Click Refresh.</p></div>`;return}

    // Futures
    let h='<div class="card"><h2>Futures & Macro Indicators</h2>';
    h+=`<p style="margin-bottom:12px">Verdict: <span class="badge ${d.futures_verdict==='BULLISH'?'bull':'neutral'}">${d.futures_verdict}</span></p>`;
    h+='<table><tr><th>Index</th><th>Level</th><th>Change</th><th>%</th><th>Signal</th></tr>';
    (d.futures||[]).forEach(f=>{
      const c=f.pct>=0?'pos':'neg';
      const s=f.signal.includes('BULL')||f.signal.includes('DECREASING')?'bull':f.signal.includes('BEAR')||f.signal.includes('INCREASING')?'bear':'neutral';
      h+=`<tr><td style="text-align:left">${f.name}</td><td>${f.level.toLocaleString()}</td><td class="${c}">${f.change>=0?'+':''}${f.change}</td><td class="${c}">${f.pct>=0?'+':''}${f.pct}%</td><td><span class="badge ${s}">${f.signal}</span></td></tr>`;
    });
    h+='</table></div>';
    document.getElementById('futures').innerHTML=h;

    // Swing trades
    h='<div class="card"><h2>Top 10 Swing Trade Opportunities</h2>';
    h+='<table><tr><th>#</th><th>Ticker</th><th>Price</th><th>Daily %</th><th>3-Mo %</th><th>Mkt Cap</th><th>Swing Score</th></tr>';
    (d.stocks||[]).forEach((s,i)=>{
      const t=i<3?' class="top3"':'';
      const dc=s.daily_pct>=0?'pos':'neg';
      const mc=s.three_mo_pct>=0?'pos':'neg';
      h+=`<tr${t}><td>${i+1}</td><td><strong>${s.ticker}</strong></td><td>$${s.price.toFixed(2)}</td><td class="${dc}">${(s.daily_pct*100).toFixed(2)}%</td><td class="${mc}">${(s.three_mo_pct*100).toFixed(2)}%</td><td>$${s.mkt_cap.toFixed(1)}B</td><td><strong>${s.swing_score}</strong></td></tr>`;
    });
    h+='</table></div>';
    document.getElementById('swing').innerHTML=h;

    // Buy Signals
    h='<div class="card"><h2>Technical Buy Signal Analysis</h2>';
    h+='<table><tr><th>Ticker</th><th>Price</th><th>Buy Score</th><th>Rating</th><th>RSI</th><th>MACD Cross</th><th>BB Pos</th><th>Above 50-SMA</th><th>Golden Cross</th><th>Vol Surge</th><th>Key Reasons</th></tr>';
    (d.buy_signals||[]).forEach(s=>{
      const rc=s.rating==='BUY'?'buy-row':s.rating==='WATCH'?'watch-row':'avoid-row';
      const bc=s.rating==='BUY'?'buy-badge':s.rating==='WATCH'?'watch-badge':'avoid-badge';
      h+=`<tr class="${rc}"><td><strong>${s.ticker}</strong></td><td>$${s.price.toFixed(2)}</td><td><strong>${s.buy_score}</strong></td><td><span class="badge ${bc}">${s.rating}</span></td><td>${s.rsi}</td><td>${s.macd_cross}</td><td>${s.bb_position}%</td><td>${s.above_50sma}</td><td>${s.golden_cross}</td><td>${s.vol_surge}</td><td style="text-align:left;font-size:12px">${s.reasons}</td></tr>`;
    });
    h+='</table></div>';
    document.getElementById('buy').innerHTML=h;

    // Technical Charts
    h='<div class="card"><h2>Technical Analysis Charts — 6-Month</h2><p style="margin-bottom:16px;color:#8b949e">Price + Bollinger Bands + 50/200 SMA | Volume | RSI (14) | MACD</p>';
    if(d.chart_tickers && d.chart_tickers.length>0){
      d.chart_tickers.forEach(t=>{
        h+=`<div style="margin-bottom:24px;text-align:center"><img src="/chart/${t}" style="max-width:100%;border:1px solid #30363d;border-radius:8px" alt="${t} chart" loading="lazy"></div>`;
      });
    } else {
      h+='<p>No charts available yet. Click Refresh to generate.</p>';
    }
    h+='</div>';
    document.getElementById('charts').innerHTML=h;

    document.getElementById('timestamp').innerText='Last updated: '+(d.generated_at||'N/A');
  }catch(e){document.getElementById('futures').innerHTML='<div class="card"><p>Could not load data. Click Refresh.</p></div>';}
}
document.getElementById('refreshBtn').addEventListener('click',async(e)=>{
  e.preventDefault();
  document.getElementById('status').innerText='Refreshing data... this takes 30-60 seconds...';
  const r=await fetch('/refresh');const d=await r.json();
  document.getElementById('status').innerText=d.status==='ok'?'Refresh complete!':'Error: '+d.error;
  setTimeout(loadDashboard,1000);
});
loadDashboard();
setInterval(loadDashboard,300000);
</script>
</body>
</html>
"""

@app.route("/")
def index():
    return render_template_string(DASHBOARD_HTML)

@app.route("/api/data")
def api_data():
    if latest["error"] and not latest["file"]:
        return jsonify({"error": latest["error"]})
    if not latest["file"]:
        return jsonify({"error": "No data yet. Click Refresh to generate."})
    summary = latest["summary"]
    stocks = []
    try:
        df = pd.read_excel(os.path.join(CONFIG["OUTPUT_DIR"], "latest.xlsx"),
                           sheet_name="Top 10 Swing Trades", skiprows=2)
        for _, r in df.iterrows():
            stocks.append({"ticker": r.get("Ticker",""), "price": float(r.get("Current Price",0)),
                           "daily_pct": float(r.get("Daily % Chg",0)),
                           "three_mo_pct": float(r.get("3-Month % Chg",0)),
                           "mkt_cap": float(r.get("Market Cap ($B)",0)),
                           "swing_score": int(r.get("Swing Score",0))})
    except Exception:
        stocks = []

    buy_sigs = []
    for s in latest.get("buy_signals", []):
        buy_sigs.append({
            "ticker": s["Ticker"], "price": s["Price"],
            "buy_score": s["Buy Score"], "rating": s["Rating"],
            "rsi": s["RSI (14)"], "macd_cross": s["MACD Cross"],
            "bb_position": s["BB Position %"],
            "above_50sma": s["Above 50-SMA"], "golden_cross": s["Golden Cross"],
            "vol_surge": s["Vol Surge"], "reasons": s["Key Reasons"],
        })

    return jsonify({
        "generated_at": latest["generated_at"],
        "futures_verdict": summary.get("futures_verdict", "N/A"),
        "futures": summary.get("futures", []),
        "stocks": stocks, "buy_signals": buy_sigs,
        "chart_tickers": list(latest.get("chart_images", {}).keys()),
    })

@app.route("/refresh")
def refresh():
    try:
        run_job(skip_email="--no-email" in sys.argv)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)})

@app.route("/chart/<ticker>")
def chart_image(ticker):
    ticker = ticker.upper()
    chart_path = os.path.join(CONFIG["OUTPUT_DIR"], "charts", f"{ticker}_chart.png")
    if os.path.exists(chart_path):
        return send_file(chart_path, mimetype="image/png")
    return "Chart not found", 404

@app.route("/download")
def download():
    path = os.path.join(CONFIG["OUTPUT_DIR"], "latest.xlsx")
    if os.path.exists(path):
        return send_file(path, as_attachment=True,
                         download_name=f"swing_dashboard_{datetime.now(CST).strftime('%Y%m%d')}.xlsx")
    return "No file yet. Visit /refresh first.", 404


# ===========================================================================
#  SCHEDULER
# ===========================================================================

def start_scheduler(skip_email=False):
    sched = BackgroundScheduler(timezone=CONFIG["TZ"])
    fn = lambda: run_job(skip_email=skip_email)
    sched.add_job(fn, CronTrigger(hour=7, minute=0), id="morning")
    sched.add_job(fn, CronTrigger(hour=9, minute=30), id="premarket")
    sched.add_job(fn, CronTrigger(hour=12, minute=0), id="midday")
    sched.add_job(fn, CronTrigger(hour=14, minute=45), id="afternoon")
    sched.start()
    log.info("Scheduler started: 7:00, 9:30, 12:00, 14:45 CST")
    return sched


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--run-once", action="store_true")
    parser.add_argument("--web-only", action="store_true")
    parser.add_argument("--no-email", action="store_true")
    args = parser.parse_args()

    if args.run_once:
        run_job(skip_email=args.no_email)
        return

    log.info("Running initial data fetch...")
    threading.Thread(target=run_job, kwargs={"skip_email": args.no_email}, daemon=True).start()
    if not args.web_only:
        start_scheduler(skip_email=args.no_email)
    log.info("Starting web dashboard on port %d", CONFIG["PORT"])
    app.run(host="0.0.0.0", port=CONFIG["PORT"], debug=False)


if __name__ == "__main__":
    main()
